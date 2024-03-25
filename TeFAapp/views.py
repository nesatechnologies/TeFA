from django.shortcuts import render, redirect
from .models import *

from datetime import datetime
from django.contrib import auth,messages

from .decorators import session_login_required

import csv
from django.db.models import Q
import openpyxl
from openpyxl import Workbook
from django.http import HttpResponse
from django.db.models import Max

from openpyxl.styles import Font, Color, Fill
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

from django.template.loader import render_to_string

from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
import pandas as pd
import io
from django.http import JsonResponse

# Create your views here.
@session_login_required
def home(request):
    if 'username' in request.session:
        data = Lead.objects.filter(status=0).order_by('-id')
        ## pagination part
        paginator = Paginator(data, 50)  # Show 10 items per page
        page = request.GET.get('page')
        try:
            data = paginator.page(page)
        except PageNotAnInteger:
            # If page is not an integer, deliver first page.
            data = paginator.page(1)
        except EmptyPage:
            # If page is out of range, deliver last page of results.
            data = paginator.page(paginator.num_pages)
            
        no_contact = Lead.objects.all().count()
        wait_call = Lead.objects.filter(status=0).count()
        conformed = Lead.objects.filter(status=1).count()
        need_following = Lead.objects.filter(status=2).count()
        denied = Lead.objects.filter(status=3).count()
        return render(request, 'home.html',{'data':data, 'wait_call':wait_call, 'no_contact':no_contact,
                                            'conformed':conformed, 'need_following':need_following, 'denied':denied})
    else:
        return redirect('/')
def homeviewall(request):
    if 'username' in request.session:
        data = Lead.objects.filter(status=0).order_by('-id')
            
        no_contact = Lead.objects.all().count()
        wait_call = Lead.objects.filter(status=0).count()
        conformed = Lead.objects.filter(status=1).count()
        need_following = Lead.objects.filter(status=2).count()
        denied = Lead.objects.filter(status=3).count()
        return render(request, 'home.html',{'data':data, 'wait_call':wait_call, 'no_contact':no_contact,
                                            'conformed':conformed, 'need_following':need_following, 'denied':denied})
    else:
        return redirect('/')
def conformed(request):
    if 'username' in request.session:
        data = Calldetails.objects.filter(lead__status=1).order_by('-id')
        ## pagination part
        paginator = Paginator(data, 50)  # Show 10 items per page
        page = request.GET.get('page')
        try:
            data = paginator.page(page)
        except PageNotAnInteger:
            # If page is not an integer, deliver first page.
            data = paginator.page(1)
        except EmptyPage:
            # If page is out of range, deliver last page of results.
            data = paginator.page(paginator.num_pages)
        return render(request, 'conformed.html', {'data':data})
    else:
        return redirect('/')
def conformedviewall(request):
    if 'username' in request.session:
        data = Calldetails.objects.filter(lead__status=1).order_by('-id')
        return render(request, 'conformed.html', {'data':data})
    else:
        return redirect('/')
def need_following(request):
    if 'username' in request.session:
        data = Calldetails.objects.filter(lead__status=2).order_by('-id')

        ## pagination part
        paginator = Paginator(data, 50)  # Show 10 items per page
        page = request.GET.get('page')
        try:
            data = paginator.page(page)
        except PageNotAnInteger:
            # If page is not an integer, deliver first page.
            data = paginator.page(1)
        except EmptyPage:
            # If page is out of range, deliver last page of results.
            data = paginator.page(paginator.num_pages)

        return render(request, 'need_following.html', {'data':data})
    else:
        return redirect('/')
def need_followingseeall(request):
    if 'username' in request.session:
        data = Calldetails.objects.filter(lead__status=2).order_by('-id')
        return render(request, 'need_following.html', {'data':data})
    else:
        return redirect('/')
def priorityonBtn(request):
    if 'username' in request.session:
        data = Calldetails.objects.filter(lead__priority=1).order_by('-id')
        data1 = Calldetails.objects.filter(lead__priority=0).order_by('-id')
        return render(request, 'need_following.html', {'data':data,'data1':data1})
    else:
        return redirect('/')
def denied(request):
    if 'username' in request.session:
        data = Calldetails.objects.filter(lead__status=3).order_by('-id')
        ## pagination part
        paginator = Paginator(data, 50)  # Show 10 items per page
        page = request.GET.get('page')
        try:
            data = paginator.page(page)
        except PageNotAnInteger:
            # If page is not an integer, deliver first page.
            data = paginator.page(1)
        except EmptyPage:
            # If page is out of range, deliver last page of results.
            data = paginator.page(paginator.num_pages)
        return render(request, 'denied.html', {'data':data})
    else:
        return redirect('/')
def deniedviewall(request):
    if 'username' in request.session:
        data = Calldetails.objects.filter(lead__status=3).order_by('-id')
        return render(request, 'denied.html', {'data':data})
    else:
        return redirect('/')
def add_customer(request):
    if 'username' in request.session:
        # inputing user data from employee side form
        if request.method == 'POST':
            phone_no = request.POST.get('phone_no')
            name = request.POST.get('name')
            course_types = request.POST.get('coursemode')
            if course_types == "Not mentioned":
                course_type = ""
            else:
                course_type = course_types
            course = request.POST.get('course')
            email = request.POST.get('email')
            place = request.POST.get('place')
            lead_date = request.POST.get('lead_date')
            remark = request.POST.get('remark')
            source = request.POST.get('source')
            degree = request.POST.get('degree')

            ######  control_no part operations ######
            # Get the last row based on the primary key
            last_row = Lead.objects.last()
            if last_row:
                # Access attributes of the last row
                print("Last row:", last_row.control_no)
                control_no = last_row.control_no
                control_no += 1
            else:
                ##### if table is empty give default starting value is given here...
                print("Table is empty")
                control_no = 5000


            ######### starting lead_no part ###########
            ## using entered 'lead_date' can used to create special type of code for "lead_no".
            ## month-year-Lnumber eg : feb-24-L1


            ### month first three letter taking part
            # Parse the date string
            date_string = lead_date
            date_object = datetime.strptime(date_string, "%Y-%m-%d")
            # Get the English month name first three letters using %b
            english_month = date_object.strftime("%b")

            #### year last two digit taking part
            # Parse the date string
            date_string = lead_date
            date_object = datetime.strptime(date_string, "%Y-%m-%d")
            # Get the last two digits of the year
            last_two_digits_year = date_object.strftime("%y")

            ### retrieve last updated row from data base and compire to done operation
            if last_row:
                # Access "lead date" attributes of the last row.
                lead_given_date1 =last_row.lead_given_date

                # Parse the date strings into datetime objects
                lead_given_date1_parse = datetime.strptime(str(lead_given_date1), "%Y-%m-%d")
                lead_date_parse = datetime.strptime(lead_date, "%Y-%m-%d")

                ## comparing the previous row lead_date and new entered one lead_date
                if lead_given_date1_parse == lead_date_parse:
                    # in the case of same date old one same lead_no is giving
                    lead_no = last_row.lead_no
                else:

                    try:
                        ### check data before entered if lend date already entered take same lead_no( not only checking just previous row )
                        lended_date_details = Lead.objects.get(lead_given_date = lead_date)
                        lead_no = lended_date_details.lead_no

                    except:
                        # split to a list to compire its month and year
                        x = str(lead_given_date1_parse).split("-")
                        y = str(lead_date_parse).split("-")
                        # in the case of different year or month compired to previous one just set val =1
                        if x[0] != y[0] or x[1] != y[1]:
                            val = 1
                            lead_no = english_month + '-' + last_two_digits_year + '-' + 'L' + str(val)
                        else:
                            # in the case of previous and new lend date month & year same
                            bfore_lead_no= last_row.lead_no
                            # taking number from previous lend no from last position and add 1 to it
                            lead_array = bfore_lead_no.split('-')
                            lead_spe_val = lead_array[-1]
                            a = lead_spe_val[1:]
                            val = int(a) + 1
                            lead_no = english_month + '-' + last_two_digits_year + '-' + 'L' + str(val)
            else:
                val = 1
                lead_no = english_month + '-' + last_two_digits_year + '-' + 'L' + str(val)

            # lead_no =english_month+'-'+last_two_digits_year+'-'+'L'+str(val)

            data = Lead(lead_given_date=lead_date, name=name, course=course, phone_no=phone_no, email=email, place=place, remark=remark, control_no=control_no, lead_no=lead_no, source=source, degree=degree, course_type=course_type)
            data.save()
            return redirect('home')
        coursedata = Courses.objects.all()
        return render(request, 'add_customer.html',{'coursedata':coursedata})
    else:
        return redirect('/')

def delete(request, id):
    if request.method == 'POST':
        data = Lead.objects.get(id=id)
        data.delete()
        return redirect('home')
    data = Lead.objects.filter(id=id)
    return render(request, 'delete.html',{'data':data})

def delete2(request, id):
    if request.method == 'POST':
        data = Lead.objects.get(id=id)
        data.delete()
        return redirect('contactbook')
    data = Lead.objects.filter(id=id)
    return render(request, 'delete2.html',{'data':data})

def delete3(request, id):
    if request.method == 'POST':
        data = Lead.objects.get(id=id)
        data.delete()
        return redirect('/')
    data = Lead.objects.filter(id=id)
    return render(request, 'delete3.html',{'data':data})

def login(request):
    if request.method == 'POST':
        print("login part")
        username = request.POST['username']
        password = request.POST['password']
        if username != '' and password != '':
            if Employee_details.objects.filter(user_name=username, password=password).exists():
                data = Employee_details.objects.filter(user_name=username, password=password).values('name', 'emp_id', 'id').first()
                print(data)
                request.session['name'] = data['name']
                request.session['emp_id'] = data['emp_id']
                request.session['username'] = username
                request.session['password'] = password
                request.session['uid'] = data['id']
                return redirect('home')
            else:
                messages.info(request, "enter valid inputs")
                return redirect('login')
        else:
            messages.info(request, "enter all inputs")
            return redirect('login')
    return render(request, 'login.html')

# def register(request):
#     if request.method == 'POST':
#         username = request.POST['username']
#         name = request.POST['name']
#         empid = request.POST['empid']
#         password = request.POST['password']
#         cpassword = request.POST['cpassword']
#         if username != '' and password != '' and cpassword !='' and empid !='' and name !='':
#             if password == cpassword:
#                 if Employee_details.objects.filter(user_name=username).exists():
#                     messages.info(request, "username is Already taken")
#                 elif Employee_details.objects.filter(emp_id=empid).exists():
#                     messages.info(request, "Employee id is Already taken")
#                 else:
#                     user = Employee_details(user_name=username,password=password,name=name,emp_id=empid)
#                     user.save()
#                     messages.info(request, "login please")
#                     return redirect('/')
#             else:
#                 messages.info(request, "passwords not matched")
#                 return redirect('register')
#         else:
#             messages.info(request, "enter all inputs")
#             return redirect('register')
#     return render(request, 'register.html')

def logout(request):
    del request.session['name']
    del request.session['emp_id']
    del request.session['username']
    del request.session['password']
    del request.session['uid']
    return redirect('login')

def call(request,id):
    if 'username' in request.session:
        if request.method == 'POST':
            selected_value = request.POST['name']
            if selected_value:
                name, emp_id = selected_value.split('|')
                # Now you have name and emp_id separately
                # Do whatever you want with these values
            else:
                # Handle case when no option is selected
                pass
            status = request.POST['status']
            name  = request.POST['name1']
            course = request.POST['course']
            phone_no = request.POST['phone_no']
            email = request.POST['email']
            place = request.POST['place']
            degree = request.POST['degree']
            course_types = request.POST.get('coursemode')
            if course_types == "Not mentioned":
                course_type = ""
            else:
                course_type = course_types

            called_meadium = request.POST['called_meadium']
            emp_remark = request.POST['remark']
            lead = Lead.objects.get(id=id)
            calls_made= Employee_details.objects.get(emp_id=emp_id)
            calls_updated_id = request.session.get('uid')
            calls_updated= Employee_details.objects.get(id=calls_updated_id)

            userdata = Calldetails(lead=lead, calls_made=calls_made, emp_remark=emp_remark, called_meadium=called_meadium, calls_updated=calls_updated)
            userdata.save()
            lead.status = status
            lead.name = name
            lead.course = course
            lead.phone_no = phone_no
            lead.email = email
            lead.place = place
            lead.degree = degree
            lead.course_type =course_type
            lead.save()
            # Redirect to the home page with refresh parameter
            return redirect('/')
        data = Lead.objects.filter(id=id)
        data1 = Employee_details.objects.all()
        coursedata = Courses.objects.all()
        return render(request, 'call.html',{'data':data,'data1':data1,'coursedata':coursedata})
    else:
        return redirect('/')

def followup(request, id):
    if 'username' in request.session:
        print("followup")
        if request.method == 'POST':
            print("followup1")
            selected_value = request.POST['name']
            if selected_value:
                name, emp_id = selected_value.split('|')
                # Now you have name and emp_id separately
                # Do whatever you want with these values
            else:
                # Handle case when no option is selected
                pass
            status = request.POST['status']

            name = request.POST['name1']
            course = request.POST['course']
            phone_no = request.POST['phone_no']
            email = request.POST['email']
            place = request.POST['place']
            degree = request.POST['degree']

            course_types = request.POST.get('coursemode')
            if course_types == "Not mentioned":
                course_type = ""
            else:
                course_type = course_types

            called_meadium = request.POST['called_meadium']
            remark = request.POST['remark']

            calldetails = Calldetails.objects.get(id=id)
            calls_made= Employee_details.objects.get(emp_id=emp_id)
            calls_updated_id = request.session.get('uid')
            calls_updated= Employee_details.objects.get(id=calls_updated_id)


            userdata = Folloup(calldetails=calldetails, remark=remark, called_meadium=called_meadium, calls_made=calls_made, calls_updated=calls_updated)
            userdata.save()

            current_followups = calldetails.no_of_followups
            calldetails.no_of_followups = current_followups + 1
            calldetails.save()


            calldetails.lead.name = name
            calldetails.lead.course = course
            calldetails.lead.course_type = course_type
            calldetails.lead.phone_no = phone_no
            calldetails.lead.email = email
            calldetails.lead.place = place
            calldetails.lead.degree = degree
            calldetails.lead.status = status
            calldetails.lead.save()
            return redirect('/')

        #### need folloup page - calldeatils id is get from call #####
        data = Calldetails.objects.filter(id=id)
        data1 = Employee_details.objects.all()
        data2 = Folloup.objects.filter(calldetails__id=id)
        data3 = Calldetails.objects.get(id=id)
        coursedata = Courses.objects.all()
        return render(request, 'folloup.html',{'data':data,'data1':data1,'data2':data2,'data3':data3,'coursedata':coursedata})
    else:
        return redirect('/')

# contactbook folloup print form id get is lead id
def followup2(request, id):
    if 'username' in request.session:
        print("followup")
        if request.method == 'POST':
            print("followup1")
            selected_value = request.POST['name']
            if selected_value:
                name, emp_id = selected_value.split('|')
                # Now you have name and emp_id separately
                # Do whatever you want with these values
            else:
                # Handle case when no option is selected
                pass
            status = request.POST['status']

            name = request.POST['name1']
            course = request.POST['course']
            phone_no = request.POST['phone_no']
            email = request.POST['email']
            place = request.POST['place']
            degree = request.POST['degree']

            course_types = request.POST.get('course_type')
            if course_types == "Not mentioned":
                course_type = ""
            else:
                course_type = course_types

            called_meadium = request.POST['called_meadium']
            remark = request.POST['remark']

            calldetails = Calldetails.objects.get(id=id)
            calls_made = Employee_details.objects.get(emp_id=emp_id)
            calls_updated_id = request.session.get('uid')
            calls_updated = Employee_details.objects.get(id=calls_updated_id)

            userdata = Folloup(calldetails=calldetails, remark=remark, called_meadium=called_meadium,
                               calls_made=calls_made, calls_updated=calls_updated)
            userdata.save()

            current_followups = calldetails.no_of_followups
            calldetails.no_of_followups = current_followups + 1
            calldetails.save()

            calldetails.lead.name = name
            calldetails.lead.course = course
            calldetails.lead.phone_no = phone_no
            calldetails.lead.email = email
            calldetails.lead.place = place
            calldetails.lead.degree = degree
            calldetails.lead.status = status
            calldetails.lead.course_type = course_type
            calldetails.lead.save()
            return redirect('/')

        #### contactbook page - lead id is get from contactbook page #####
        calldet_id = Calldetails.objects.get(lead__id = id)
        calldetails_id = calldet_id.id
        print("1")
        data = Calldetails.objects.filter(id=calldetails_id)
        print("2")
        data1 = Employee_details.objects.all()
        print("3")
        data2 = Folloup.objects.filter(calldetails__id=calldetails_id)
        print("4")
        data3 = Calldetails.objects.get(id=calldetails_id)
        print("5")
        coursedata = Courses.objects.all()
        print("6")
        return render(request, 'folloup_2.html',
                      {'data': data, 'data1': data1, 'data2': data2, 'data3': data3, 'coursedata': coursedata})
    else:
        return redirect('/')

def followup_actions(request,id):
    if 'username' in request.session:
        data = Folloup.objects.filter(calldetails__id = id)
        data1 = Calldetails.objects.get(id=id)
        return render(request,'followup_actions.html',{'data':data,'data1':data1})
    else:
        return redirect('/')

def upload_csv(request):
    message = ""  # Default message
    csv_data = None  # Default CSV data
    if request.method == 'POST' and request.FILES.get('xlsx_file'):
        doc_files = request.FILES['xlsx_file']

        if doc_files.name.endswith('.xlsx'):
            df = pd.read_excel(doc_files)
            csv_file = df.to_csv(index=False)
        # elif doc_files.name.endswith('.csv'):
        #     csv_file = doc_files
            # Process the uploaded CSV file
            try:
                # Decode and process the CSV file
                csv_data = io.StringIO(csv_file)
                csv_data = csv.reader(csv_data)
                for row in csv_data:
                    # Process each row of the CSV file
                    if row[0] == 'SL.NO':
                        continue
                    else:
                        print(row)
                        control_no = int(row[1])
                        print("1")
                        lead_no = str(row[4])
                        print("2")

                        # Input date string
                        date_string = str(row[3])
                        print("3")
                        print("***************")
                        print(date_string)
                        print("***************")
                        my_string = date_string
                        # Split the string by comma
                        try:
                            print("before split")
                            split_values = my_string.split(' ')
                            first_value = split_values[0]
                            print("after split")
                            
                        except:
                            first_value = my_string
                            print("")

                        # Get the first value from the list
                        
                        print("_____first_value_____")
                        print(first_value)


                        # Parse the date string into a datetime object
                        formats_to_check = ["%d/%m/%y", "%d/%m/%Y", "%d-%m-%y", "%d-%m-%Y", "%Y-%m-%d", "%y-%m-%d"]
                        for date_format in formats_to_check:
                            try:
                                # Attempt to parse the date string using the current format
                                date_object = datetime.strptime(first_value, date_format)
                            except ValueError:
                                # If parsing fails, continue to the next format
                                continue
                    

                        # Format the datetime object in the desired format
                        date_part = date_object.date()
                        formatted_date = date_part.strftime("%Y-%m-%d")
                        print(formatted_date)
                        lead_given_date = formatted_date
                        print(lead_given_date)
                        print("4")

                        source = str(row[12])
                        name = str(row[5])
                        print("5")
                        phone_no = int(row[8])
                        email = (row[9])
                        print("6")
                        place = str(row[10])
                        degree = str(row[13])
                        course_type = str(row[6])
                        course = str(row[7])
                        remark = str(row[11])
                        print("7")
                        data = Lead(lead_given_date=lead_given_date, name=name, course=course, phone_no=phone_no, email=email,
                                    place=place, remark=remark, control_no=control_no, lead_no=lead_no, source=source,
                                    degree=degree, course_type=course_type)
                        data.save()
                        print("8")
                        rowstat = str(row[14])
                        if rowstat == "wait for call":
                            continue

                        ##### initial call part
                        initial_call = str(row[19])
                        print(initial_call)
                        calls_made= Employee_details.objects.get(user_name=initial_call)

                        calls_updated_id = request.session.get('uid')
                        calls_updated = Employee_details.objects.get(id=calls_updated_id)




                        print("*******date string********")
                        # Input date string
                        date_string2 = str(row[17])
                        print("9")
                        print("***************")
                        print(date_string2)
                        print("***************")
                        
                        print("10")
                        my_string2 = date_string2
                        # Split the string by comma
                        try:
                            print("before split2")
                            split_values = my_string2.split(' ')
                            first_value2 = split_values[0]
                            print("after split2")
                            
                        except:
                            first_value2 = my_string2
                            print("")

                        # Get the first value from the list
                        
                        print("_____first_value2_____")
                        print(first_value2)
                        # Parse the date string into a datetime object
                        formats_to_check = ["%d/%m/%y", "%d/%m/%Y", "%d-%m-%y", "%d-%m-%Y", "%Y-%m-%d", "%y-%m-%d"]
                        for date_format in formats_to_check:
                            try:
                                # Attempt to parse the date string using the current format
                                date_object2 = datetime.strptime(first_value2, date_format)
                            except ValueError:
                                # If parsing fails, continue to the next format
                                continue
                            

                        # Format the datetime object in the desired format
                        date_part2 = date_object2.date()
                        formatted_date2 = date_part2.strftime("%Y-%m-%d")
                        print(formatted_date2)
                        initial_call_date = formatted_date2
                        print("11")


                        initial_call_remark= str(row[16])
                        called_meadium=str(row[18])


                        lead = Lead.objects.get(id=data.id)
                        stat = str(row[14])
                        if stat == "conformed":
                            print("conformed")
                            lead.status = 1
                        elif stat == "need following":
                            print("need following")
                            lead.status = 2
                        elif stat == "denied":
                            print("denied")
                            lead.status = 3
                        lead.save()

                        data2= Calldetails(lead=lead, calls_made=calls_made, emp_remark=initial_call_remark, called_datetime=initial_call_date, called_meadium=called_meadium, calls_updated=calls_updated)
                        data2.save()

                        # need following part decument update
                        new_data = row[21:]
                        print(new_data)
                        sublist = []
                        for i, item in enumerate(new_data):
                            print(item)
                            # Skip every fourth position containing an empty string
                            if (i + 1) % 4 == 0 and item == "":
                                continue
                            # Append the current item to the sublist
                            sublist.append(item)
                            if len(sublist) == 3:
                                if sublist[0] == "" and sublist[1] == "" and sublist[2] == "":
                                    continue
                                else:
                                    print(sublist)
                                    calldetails = Calldetails.objects.get(id=data2.id)
                                    print("1.1")
                                    remark = sublist[2]
                                    print("1.2")
                                    calls_made = Employee_details.objects.get(user_name=sublist[0])
                                    print("1.3")
                                    calls_updated_id = request.session.get('uid')
                                    print("1.4")
                                    calls_updated = Employee_details.objects.get(id=calls_updated_id)
                                    print("1.5")
                                    called_meadium = ""
                                    print("1.6")

                                    if sublist[1] == "":
                                        data3 = Folloup(calldetails=calldetails, remark=remark, calls_made=calls_made,
                                                        called_meadium=called_meadium, calls_updated=calls_updated)
                                        data3.save()
                                        print("2")
                                    else:
                                        print("3")
                                        print(sublist[1])
                                        print(type(sublist[1]))
                                        my_string3 = sublist[1]
                                        # Split the string by comma
                                        try:
                                            print("before split3")
                                            split_values = my_string3.split(' ')
                                            first_value3 = split_values[0]
                                            print("after split3")
                                            
                                        except:
                                            first_value3 = my_string3
                                            print("")

                                        # Get the first value from the list
                                        
                                        print("_____first_value3_____")
                                        print(first_value3)

                                        # Parse the date string into a datetime object
                                        formats_to_check = ["%d/%m/%y", "%d/%m/%Y", "%d-%m-%y", "%d-%m-%Y", "%Y-%m-%d", "%y-%m-%d"]
                                        for date_format in formats_to_check:
                                            try:
                                                # Attempt to parse the date string using the current format
                                                date_object3 = datetime.strptime(first_value3, date_format)
                                            except ValueError:
                                                # If parsing fails, continue to the next format
                                                continue
                                    
                                        print("4-followup")
                                    
                                        print(date_object3)
                                       
                                        date_part3 = date_object3.date()
                                        print(date_part3)
                                        formatted_date = date_part3.strftime("%Y-%m-%d")
                                        called_datetime = formatted_date
                                        print("5")
                                        data3 = Folloup(calldetails=calldetails, remark=remark, calls_made=calls_made,
                                                        called_datetime=called_datetime, called_meadium=called_meadium,
                                                        calls_updated=calls_updated)
                                        data3.save()

                                    sublist = []
                                    current_followups = calldetails.no_of_followups
                                    calldetails.no_of_followups = current_followups + 1
                                    calldetails.save()


                print("CSV file uploaded and processed successfully.")
                message = "CSV file uploaded and processed successfully."
            except Exception as e:
                message = f"Error processing CSV file: {e}"
        else:
            message = "Please upload a valid CSV file."
    else:
        message = "No file uploaded."
    return redirect('/')

def contactbook(request):
    if 'username' in request.session:
        data = Lead.objects.all().order_by('-control_no')
        ## pagination part
        paginator = Paginator(data, 100)  # Show 10 items per page
        page = request.GET.get('page')
        try:
            data = paginator.page(page)
        except PageNotAnInteger:
            # If page is not an integer, deliver first page.
            data = paginator.page(1)
        except EmptyPage:
            # If page is out of range, deliver last page of results.
            data = paginator.page(paginator.num_pages)
        return render(request,'contactbook.html',{'data':data})
    else:
        return redirect('/')
    
def contactbookviewall(request):
    if 'username' in request.session:
        data = Lead.objects.all().order_by('-control_no')
        return render(request,'contactbook.html',{'data':data})
    else:
        return redirect('/')
    
def searchresult(request):
    if 'username' in request.session:
        try:
            products= None
            query= None
            if 'q' in request.GET:
                query = request.GET.get('q')
                products= Lead.objects.all().filter(Q(control_no__contains = query) | Q(date_time_added__contains = query) | Q(lead_given_date__contains = query) | Q(lead_no__contains = query) | Q(name__contains = query) | Q(course__contains = query) | Q(phone_no__contains = query) | Q(email__contains = query) | Q(place__contains = query) | Q(remark__contains = query) | Q(status__contains = query) | Q(source__contains = query) | Q(degree__contains = query))
                for m in products:
                    data = Calldetails.objects.filter(lead__id = m.id)
                return render(request, 'search.html', {'query':query, 'products':products, 'data':data})
        except:
            if 'q' in request.GET:
                query = request.GET.get('q')
            return render(request, 'search.html', {'query':query})
    else:
        return redirect('/')

# CONTACTBOOK
def export_to_excel(request):
    if 'username' in request.session:
        # Create a new workbook
        wb = Workbook()

        # Activate the first sheet
        ws = wb.active
        ws.title = "Contactbook Data"

         # Write headers
        # headers = ["Control No", "Date Time Added", "Lead Given Date", "Lead No", "Name", "Course", "Phone No", "Email", "Place", "Remark", "Status", "Source", "Degree"]
        # ws.append(headers)

        # Define a red fill style
        yellow_fill = PatternFill(start_color='fef2cb', end_color='fef2cb', fill_type='solid')

        headers = ["Control No", "Date Added", "Lead Given Date", "Lead No", "Name", "Course type", "Course", "Phone No", "Email",
                   "Place", "Remark", "Source", "Qualification", "Status"]

        # Write headers with the red fill style
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = yellow_fill

        # Fetch data from Lead model
        leads = Lead.objects.all()

        # Write data rows
        for lead in leads:
            statusval = ""
            if lead.status == 0:
                statusval = "wait for call"
            elif lead.status == 1:
                statusval = "conformed"
            elif lead.status == 2:
                statusval = "need following"
            elif lead.status == 3:
                statusval = "denied"

            row = [lead.control_no, lead.date_time_added, lead.lead_given_date, lead.lead_no, lead.name, lead.course_type, lead.course, lead.phone_no, lead.email, lead.place, lead.remark, lead.source, lead.degree, statusval]
            ws.append(row)


            ##### increase cell width ####
            for col in ws.columns:
                max_length = 0
                column = col[0].column  # Get the column index
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2  # Adjust the multiplication factor as needed
                ws.column_dimensions[get_column_letter(column)].width = adjusted_width


                # Define red fill pattern
                red_fill = PatternFill(start_color='ffcccc', end_color='ffcccc', fill_type='solid')  # Red fill

                # Get the index of the "Phone No" column
                phone_no_column_index = headers.index("Phone No") + 1  # Adding 1 because index starts from 1 in openpyxl

                # Iterate through each cell in the "Phone No" column
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=phone_no_column_index, max_col=phone_no_column_index):
                    for cell in row:
                        cell.fill = red_fill  # Apply red fill color to each cell


                # Define red fill pattern
                red_fill = PatternFill(start_color='ffcccc', end_color='ffcccc', fill_type='solid')  # Red fill
                green_fill = PatternFill(start_color='c5e0b3', end_color='c5e0b3', fill_type='solid')  # green fill
                blue_fill = PatternFill(start_color='60cbf3', end_color='60cbf3', fill_type='solid')  # blue fill
                dark_fill = PatternFill(start_color='DCDCDC', end_color='DCDCDC', fill_type='solid')  # dark fill

                # Get the index of the "Phone No" column
                status_no_column_index = headers.index("Status") + 1  # Adding 1 because index starts from 1 in openpyxl

                # Iterate through each cell in the "Phone No" column
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=status_no_column_index, max_col=status_no_column_index):
                    for cell in row:
                        if cell.value == "wait for call":
                            cell.fill = blue_fill  # Apply blue fill color to each cell
                        if cell.value == "conformed":
                            cell.fill = green_fill  # Apply green fill color to each cell
                        if cell.value == "need following":
                            cell.fill = dark_fill  # Apply dark fill color to each cell
                        if cell.value == "denied":
                            cell.fill = red_fill  # Apply red fill color to each cell

        # Create a response object
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=contactbook.xlsx'

        # Save the workbook to the response
        wb.save(response)

        return response
    else:
        return redirect('/')
    

def contactbook_detail_Report_export_to_excel(request):
    if 'username' in request.session:
        k = 0

        calldetails_data = Calldetails.objects.all()

        # Create a new workbook
        wb = Workbook()

        # Activate the first sheet
        ws = wb.active
        ws.title = "Need following Data"

        # def get_folloup_headers(calldetail):
        #     """
        #     Generates follow-up headers based on the number of follow-up entries for a calldetail.
        #     """
        #     folloup_count = calldetail.folloup_set.all().count()
        #     headers = []
        #     for _ in range(folloup_count):
        #         headers.extend(["", "Folloup Remark", "Folloup Updated", "Made By"])
        #     return headers


        def get_folloup_headers():
            #### print no of followup tables in ptint doc based on status
            largenofollow = 0
            for g in calldetails_data:
                nofollow = g.no_of_followups
                if largenofollow < nofollow:
                    largenofollow = nofollow
            highest_followups = largenofollow
            highest_followups = highest_followups
            highest_followups -= 1
            # folloup_count = calldetail.folloup_set.all().count()
            headers = []
            if highest_followups > 0:  # Add check to avoid empty headers if no follow-ups
                for i in range(highest_followups):
                    headers.extend(["", f"Follow-Up-{i+1}", "Date", "Remark"])
            return headers

        # Define base headers
        base_headers = [
            "SL.NO",
            "Control No",
            "Date Added",
            "Lead Given Date",
            "Lead No",
            "Name",
            "course type",
            "Course",
            "Phone No",
            "Email",
            "Place",
            "Lead Remark",
            "Lead Source",
            "Qualification",
            "status",
            "",
            "Initial Employee Remark",
            "Initial Called Date",
            "Source",
            "Initial Call Made",
        ]
        yellow_fill = PatternFill(start_color='fef2cb', end_color='fef2cb', fill_type='solid')  # Red fill

        # Define white fill pattern
        green_fill = PatternFill(start_color='c5e0b3', end_color='c5e0b3', fill_type='solid')  # green fill

        # Combine base headers and dynamically generated follow-up headers
        # headers = base_headers + sum([get_folloup_headers(calldetail) for calldetail in calldetails_data], [])
        headers = base_headers + sum([get_folloup_headers()], [])

        # Define font style for bold
        bold_font = Font(bold=True)
        

        # Write headers to the first row with appropriate fill applied
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header

            if cell.value:  # Check if cell value is not empty
                if col_idx <= len(base_headers):  # Apply red fill to base headers
                    cell.fill = yellow_fill
                else:  # Apply white fill to follow-up headers
                    cell.fill = green_fill

            # Apply bold font to each header cell
            cell.font = bold_font

        # Write headers to the first row only
        # Increase the height of the first row
        ws.row_dimensions[1].height = 30  # Adjust height as needed

      

        ##### add lead (wait for call details to document)
        waitforcalls = Lead.objects.filter(status=0)
         # Write data rows
        for waitforcall in waitforcalls:
            statusval = ""
            if waitforcall.status == 0:
                statusval = "wait for call"
                k = k + 1

            row = [k, waitforcall.control_no, waitforcall.date_time_added.strftime("%d-%m-%Y"), waitforcall.lead_given_date.strftime("%d-%m-%Y"), waitforcall.lead_no, waitforcall.name, waitforcall.course_type, waitforcall.course, waitforcall.phone_no, waitforcall.email, waitforcall.place, waitforcall.remark, waitforcall.source, waitforcall.degree, statusval]
            ws.append(row)


        for calldetail in calldetails_data:
            # Initialize empty lists to store folloup details
            folloup_remarks = []
            folloup_updated = []
            call_made_by = []

            # Retrieve related Folloup data
            folloup_data = Folloup.objects.filter(calldetails=calldetail)

            # Extract folloup details into separate lists
            for folloup in folloup_data:
                folloup_remarks.append(folloup.remark)
                folloup_updated.append(folloup.called_datetime.strftime("%d-%m-%Y"))  # Format date/time
                call_made_by.append(folloup.calls_made.name)  # Assuming 'name' is the relevant field


            statusval = ""
            if calldetail.lead.status == 0:
                statusval = "wait for call"
            elif calldetail.lead.status == 1:
                statusval = "conformed"
            elif calldetail.lead.status == 2:
                statusval = "need following"
            elif calldetail.lead.status == 3:
                statusval = "denied"

            # Extract Calldetails data
            row = []  # Create an empty row

            # Add lead details
            k = k + 1
            row.extend([
                k,
                calldetail.lead.control_no,
                calldetail.lead.date_time_added.strftime("%d-%m-%Y"),
                calldetail.lead.lead_given_date.strftime("%d-%m-%Y"),
                calldetail.lead.lead_no,
                calldetail.lead.name,
                calldetail.lead.course_type,
                calldetail.lead.course,
                calldetail.lead.phone_no,
                calldetail.lead.email,
                calldetail.lead.place,
                calldetail.lead.remark,
                calldetail.lead.source,
                calldetail.lead.degree,
                statusval,

                "",
                calldetail.emp_remark,
                calldetail.called_datetime.strftime("%d-%m-%Y"),
                calldetail.called_meadium,
                calldetail.calls_made.user_name,
            ])

            # Add folloup details with corresponding headings and empty columns
            for remark, updated, made_by in zip(folloup_remarks, folloup_updated, call_made_by):
                row.append("") # Add an empty column between each set of follow-up details
                row.append(made_by)
                row.append(updated)
                row.append(remark)


            # Write the combined row to the worksheet
            ws.append(row)


        ##### increase cell width ####
        for col in ws.columns:
            max_length = 0
            column = col[0].column  # Get the column index
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2  # Adjust the multiplication factor as needed
            ws.column_dimensions[get_column_letter(column)].width = adjusted_width




        # Define red fill pattern
        red_fill = PatternFill(start_color='ffcccc', end_color='ffcccc', fill_type='solid')  # Red fill

        # Get the index of the "Phone No" column
        phone_no_column_index = headers.index("Phone No") + 1  # Adding 1 because index starts from 1 in openpyxl

        # Iterate through each cell in the "Phone No" column
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=phone_no_column_index, max_col=phone_no_column_index):
            for cell in row:
                cell.fill = red_fill  # Apply red fill color to each cell


        # Define grape colour fill pattern
        grape_fill = PatternFill(start_color='ffccff', end_color='ffccff', fill_type='solid')  # grape fill

        # Get the index of the "Phone No" column
        phone_no_column_index = headers.index("Initial Call Made") + 1  # Adding 1 because index starts from 1 in openpyxl

        # Iterate through each cell in the "Phone No" column
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=phone_no_column_index, max_col=phone_no_column_index):
            for cell in row:
                cell.fill = grape_fill  # Apply red fill color to each cell

        
        # Iterate through each cell in the worksheet and set text alignment to center
        for row in ws.iter_rows():
            for cell in row:
                alignment = Alignment(horizontal='center', vertical='center')  # Center alignment
                cell.alignment = alignment

        # Define font style for Calibri
        calibri_font = Font(name='Calibri')
        # Iterate through each cell in the worksheet and set text to calibri_font
        for row in ws.iter_rows():
            for cell in row:
                # Apply Calibri font to each header cell
                cell.font = calibri_font

        # Define red fill pattern
        red_fill = PatternFill(start_color='ffcccc', end_color='ffcccc', fill_type='solid')  # Red fill
        green_fill = PatternFill(start_color='c5e0b3', end_color='c5e0b3', fill_type='solid')  # green fill
        blue_fill = PatternFill(start_color='60cbf3', end_color='60cbf3', fill_type='solid')  # blue fill
        dark_fill = PatternFill(start_color='DCDCDC', end_color='DCDCDC', fill_type='solid')  # dark fill

        # Get the index of the "Phone No" column
        status_no_column_index = headers.index("status") + 1  # Adding 1 because index starts from 1 in openpyxl

        # Iterate through each cell in the "Phone No" column
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=status_no_column_index, max_col=status_no_column_index):
            for cell in row:
                if cell.value == "wait for call":
                    cell.fill = blue_fill  # Apply blue fill color to each cell
                if cell.value == "conformed":
                    cell.fill = green_fill  # Apply green fill color to each cell
                if cell.value == "need following":
                    cell.fill = dark_fill  # Apply dark fill color to each cell
                if cell.value == "denied":
                    cell.fill = red_fill  # Apply red fill color to each cell


        # Create a response object
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Detail_Report.xlsx'

        # Save the workbook to the response
        wb.save(response)
        return response
    else:
        return redirect('/')





def need_following_export_to_excel(request):
    if 'username' in request.session:
        k = 0

        calldetails_data = Calldetails.objects.filter(lead__status=2)

        # Create a new workbook
        wb = Workbook()

        # Activate the first sheet
        ws = wb.active
        ws.title = "Need following Data"

        # def get_folloup_headers(calldetail):
        #     """
        #     Generates follow-up headers based on the number of follow-up entries for a calldetail.
        #     """
        #     folloup_count = calldetail.folloup_set.all().count()
        #     headers = []
        #     for _ in range(folloup_count):
        #         headers.extend(["", "Folloup Remark", "Folloup Updated", "Made By"])
        #     return headers


        def get_folloup_headers():
            #### print no of followup tables in ptint doc based on status
            largenofollow = 0
            for g in calldetails_data:
                nofollow = g.no_of_followups
                if largenofollow < nofollow:
                    largenofollow = nofollow
            highest_followups = largenofollow
            highest_followups = highest_followups
            highest_followups -= 1
            # folloup_count = calldetail.folloup_set.all().count()
            headers = []
            if highest_followups > 0:  # Add check to avoid empty headers if no follow-ups
                for i in range(highest_followups):
                    headers.extend(["", f"Follow-Up-{i+1}", "Date", "Remark"])
            return headers

        # Define base headers
        base_headers = [
            "SL.NO",
            "Control No",
            "Date Added",
            "Lead Given Date",
            "Lead No",
            "Name",
            "course type",
            "Course",
            "Phone No",
            "Email",
            "Place",
            "Lead Remark",
            "Lead Source",
            "Qualification",
            "status",
            "",
            "Initial Employee Remark",
            "Initial Called Date",
            "Source",
            "Initial Call Made",
        ]
        yellow_fill = PatternFill(start_color='fef2cb', end_color='fef2cb', fill_type='solid')  # Red fill

        # Define white fill pattern
        green_fill = PatternFill(start_color='c5e0b3', end_color='c5e0b3', fill_type='solid')  # green fill

        # Combine base headers and dynamically generated follow-up headers
        # headers = base_headers + sum([get_folloup_headers(calldetail) for calldetail in calldetails_data], [])
        headers = base_headers + sum([get_folloup_headers()], [])

        # Define font style for bold
        bold_font = Font(bold=True)
        

        # Write headers to the first row with appropriate fill applied
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header

            if cell.value:  # Check if cell value is not empty
                if col_idx <= len(base_headers):  # Apply red fill to base headers
                    cell.fill = yellow_fill
                else:  # Apply white fill to follow-up headers
                    cell.fill = green_fill

            # Apply bold font to each header cell
            cell.font = bold_font

        # Write headers to the first row only
        # Increase the height of the first row
        ws.row_dimensions[1].height = 30  # Adjust height as needed

        # # Remove color from empty cells
        # for row in ws.iter_rows():
        #     for cell in row:
        #         if cell.value is None:
        #             cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')  # White fill

        # ws.append(headers)
        # Set column width for each column


        for calldetail in calldetails_data:
            # Initialize empty lists to store folloup details
            folloup_remarks = []
            folloup_updated = []
            call_made_by = []

            # Retrieve related Folloup data
            folloup_data = Folloup.objects.filter(calldetails=calldetail)

            # Extract folloup details into separate lists
            for folloup in folloup_data:
                folloup_remarks.append(folloup.remark)
                folloup_updated.append(folloup.called_datetime.strftime("%d-%m-%Y"))  # Format date/time
                call_made_by.append(folloup.calls_made.name)  # Assuming 'name' is the relevant field

            # Extract Calldetails data
            row = []  # Create an empty row

            # Add lead details
            k = k + 1
            row.extend([
                k,
                calldetail.lead.control_no,
                calldetail.lead.date_time_added.strftime("%d-%m-%Y"),
                calldetail.lead.lead_given_date.strftime("%d-%m-%Y"),
                calldetail.lead.lead_no,
                calldetail.lead.name,
                calldetail.lead.course_type,
                calldetail.lead.course,
                calldetail.lead.phone_no,
                calldetail.lead.email,
                calldetail.lead.place,
                calldetail.lead.remark,
                calldetail.lead.source,
                calldetail.lead.degree,
                "need following",

                "",
                calldetail.emp_remark,
                calldetail.called_datetime.strftime("%d-%m-%Y"),
                calldetail.called_meadium,
                calldetail.calls_made.user_name,
            ])

            # Add folloup details with corresponding headings and empty columns
            for remark, updated, made_by in zip(folloup_remarks, folloup_updated, call_made_by):
                row.append("") # Add an empty column between each set of follow-up details
                row.append(made_by)
                row.append(updated)
                row.append(remark)


            # Write the combined row to the worksheet
            ws.append(row)


        ##### increase cell width ####
        for col in ws.columns:
            max_length = 0
            column = col[0].column  # Get the column index
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2  # Adjust the multiplication factor as needed
            ws.column_dimensions[get_column_letter(column)].width = adjusted_width




        # Define red fill pattern
        red_fill = PatternFill(start_color='ffcccc', end_color='ffcccc', fill_type='solid')  # Red fill

        # Get the index of the "Phone No" column
        phone_no_column_index = headers.index("Phone No") + 1  # Adding 1 because index starts from 1 in openpyxl

        # Iterate through each cell in the "Phone No" column
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=phone_no_column_index, max_col=phone_no_column_index):
            for cell in row:
                cell.fill = red_fill  # Apply red fill color to each cell


        # Define grape colour fill pattern
        grape_fill = PatternFill(start_color='ffccff', end_color='ffccff', fill_type='solid')  # grape fill

        # Get the index of the "Phone No" column
        phone_no_column_index = headers.index("Initial Call Made") + 1  # Adding 1 because index starts from 1 in openpyxl

        # Iterate through each cell in the "Phone No" column
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=phone_no_column_index, max_col=phone_no_column_index):
            for cell in row:
                cell.fill = grape_fill  # Apply red fill color to each cell

        
        # Iterate through each cell in the worksheet and set text alignment to center
        for row in ws.iter_rows():
            for cell in row:
                alignment = Alignment(horizontal='center', vertical='center')  # Center alignment
                cell.alignment = alignment

        # Define font style for Calibri
        calibri_font = Font(name='Calibri')
        # Iterate through each cell in the worksheet and set text to calibri_font
        for row in ws.iter_rows():
            for cell in row:
                # Apply Calibri font to each header cell
                cell.font = calibri_font


        # Create a response object
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=needfollowing.xlsx'

        # Save the workbook to the response
        wb.save(response)
        return response
    else:
        return redirect('/')

def conformed_export_to_excel(request):
    if 'username' in request.session:
        k = 0

        calldetails_data = Calldetails.objects.filter(lead__status=1)

        # Create a new workbook
        wb = Workbook()

        # Activate the first sheet
        ws = wb.active
        ws.title = "Need following Data"

        # def get_folloup_headers(calldetail):
        #     """
        #     Generates follow-up headers based on the number of follow-up entries for a calldetail.
        #     """
        #     folloup_count = calldetail.folloup_set.all().count()
        #     headers = []
        #     for _ in range(folloup_count):
        #         headers.extend(["", "Folloup Remark", "Folloup Updated", "Made By"])
        #     return headers


        def get_folloup_headers():
            #### print no of followup tables in ptint doc based on status
            largenofollow = 0
            for g in calldetails_data:
                nofollow = g.no_of_followups
                if largenofollow < nofollow:
                    largenofollow = nofollow

            print(largenofollow)
            highest_followups = largenofollow
            highest_followups -= 1
            # folloup_count = calldetail.folloup_set.all().count()
            headers = []
            if highest_followups > 0:  # Add check to avoid empty headers if no follow-ups
                for i in range(highest_followups):
                    headers.extend(["", f"Follow-Up-{i+1}", "Date", "Remark"])
            return headers

        # Define base headers
        base_headers = [
            "SL.NO",
            "Control No",
            "Date Added",
            "Lead Given Date",
            "Lead No",
            "Name",
            "course type",
            "Course",
            "Phone No",
            "Email",
            "Place",
            "Lead Remark",
            "Lead Source",
            "Qualification",
            "status",
            "",
            "Initial Employee Remark",
            "Initial Called Date",
            "Source",
            "Initial Call Made",
        ]
        yellow_fill = PatternFill(start_color='fef2cb', end_color='fef2cb', fill_type='solid')  # Red fill

        # Define white fill pattern
        green_fill = PatternFill(start_color='c5e0b3', end_color='c5e0b3', fill_type='solid')  # green fill

        # Combine base headers and dynamically generated follow-up headers
        # headers = base_headers + sum([get_folloup_headers(calldetail) for calldetail in calldetails_data], [])
        headers = base_headers + sum([get_folloup_headers()], [])

        # Define font style for bold
        bold_font = Font(bold=True)
        

        # Write headers to the first row with appropriate fill applied
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header

            if cell.value:  # Check if cell value is not empty
                if col_idx <= len(base_headers):  # Apply red fill to base headers
                    cell.fill = yellow_fill
                else:  # Apply white fill to follow-up headers
                    cell.fill = green_fill

            # Apply bold font to each header cell
            cell.font = bold_font

        # Write headers to the first row only
        # Increase the height of the first row
        ws.row_dimensions[1].height = 30  # Adjust height as needed

        # # Remove color from empty cells
        # for row in ws.iter_rows():
        #     for cell in row:
        #         if cell.value is None:
        #             cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')  # White fill

        # ws.append(headers)
        # Set column width for each column


        for calldetail in calldetails_data:
            # Initialize empty lists to store folloup details
            folloup_remarks = []
            folloup_updated = []
            call_made_by = []

            # Retrieve related Folloup data
            folloup_data = Folloup.objects.filter(calldetails=calldetail)

            # Extract folloup details into separate lists
            for folloup in folloup_data:
                folloup_remarks.append(folloup.remark)
                folloup_updated.append(folloup.called_datetime.strftime("%d-%m-%Y"))  # Format date/time
                call_made_by.append(folloup.calls_made.name)  # Assuming 'name' is the relevant field

            # Extract Calldetails data
            row = []  # Create an empty row

            # Add lead details
            k = k + 1
            row.extend([
                k,
                calldetail.lead.control_no,
                calldetail.lead.date_time_added.strftime("%d-%m-%Y"),
                calldetail.lead.lead_given_date.strftime("%d-%m-%Y"),
                calldetail.lead.lead_no,
                calldetail.lead.name,
                calldetail.lead.course_type,
                calldetail.lead.course,
                calldetail.lead.phone_no,
                calldetail.lead.email,
                calldetail.lead.place,
                calldetail.lead.remark,
                calldetail.lead.source,
                calldetail.lead.degree,
                "conformed",

                "",
                calldetail.emp_remark,
                calldetail.called_datetime.strftime("%d-%m-%Y"),
                calldetail.called_meadium,
                calldetail.calls_made.user_name,
            ])

            # Add folloup details with corresponding headings and empty columns
            for remark, updated, made_by in zip(folloup_remarks, folloup_updated, call_made_by):
                row.append("") # Add an empty column between each set of follow-up details
                row.append(made_by)
                row.append(updated)
                row.append(remark)


            # Write the combined row to the worksheet
            ws.append(row)


        ##### increase cell width ####
        for col in ws.columns:
            max_length = 0
            column = col[0].column  # Get the column index
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2  # Adjust the multiplication factor as needed
            ws.column_dimensions[get_column_letter(column)].width = adjusted_width




        # Define red fill pattern
        red_fill = PatternFill(start_color='ffcccc', end_color='ffcccc', fill_type='solid')  # Red fill

        # Get the index of the "Phone No" column
        phone_no_column_index = headers.index("Phone No") + 1  # Adding 1 because index starts from 1 in openpyxl

        # Iterate through each cell in the "Phone No" column
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=phone_no_column_index, max_col=phone_no_column_index):
            for cell in row:
                cell.fill = red_fill  # Apply red fill color to each cell


        # Define grape colour fill pattern
        grape_fill = PatternFill(start_color='ffccff', end_color='ffccff', fill_type='solid')  # grape fill

        # Get the index of the "Phone No" column
        phone_no_column_index = headers.index("Initial Call Made") + 1  # Adding 1 because index starts from 1 in openpyxl

        # Iterate through each cell in the "Phone No" column
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=phone_no_column_index, max_col=phone_no_column_index):
            for cell in row:
                cell.fill = grape_fill  # Apply red fill color to each cell

        
        # Iterate through each cell in the worksheet and set text alignment to center
        for row in ws.iter_rows():
            for cell in row:
                alignment = Alignment(horizontal='center', vertical='center')  # Center alignment
                cell.alignment = alignment

        # Define font style for Calibri
        calibri_font = Font(name='Calibri')
        # Iterate through each cell in the worksheet and set text to calibri_font
        for row in ws.iter_rows():
            for cell in row:
                # Apply Calibri font to each header cell
                cell.font = calibri_font


        # Create a response object
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=conformed.xlsx'

        # Save the workbook to the response
        wb.save(response)
        return response
    else:
        return redirect('/')

def denied_export_to_excel(request):
    if 'username' in request.session:
        k = 0

        calldetails_data = Calldetails.objects.filter(lead__status=3)

        # Create a new workbook
        wb = Workbook()

        # Activate the first sheet
        ws = wb.active
        ws.title = "Need following Data"

        # def get_folloup_headers(calldetail):
        #     """
        #     Generates follow-up headers based on the number of follow-up entries for a calldetail.
        #     """
        #     folloup_count = calldetail.folloup_set.all().count()
        #     headers = []
        #     for _ in range(folloup_count):
        #         headers.extend(["", "Folloup Remark", "Folloup Updated", "Made By"])
        #     return headers


        def get_folloup_headers():
            #### print no of followup tables in ptint doc based on status
            largenofollow = 0
            for g in calldetails_data:
                nofollow = g.no_of_followups
                if largenofollow < nofollow:
                    largenofollow = nofollow
            highest_followups = largenofollow
            highest_followups -= 1
            # folloup_count = calldetail.folloup_set.all().count()
            headers = []
            if highest_followups > 0:  # Add check to avoid empty headers if no follow-ups
                for i in range(highest_followups):
                    headers.extend(["", f"Follow-Up-{i+1}", "Date", "Remark"])
            return headers

        # Define base headers
        base_headers = [
            "SL.NO",
            "Control No",
            "Date Added",
            "Lead Given Date",
            "Lead No",
            "Name",
            "course type",
            "Course",
            "Phone No",
            "Email",
            "Place",
            "Lead Remark",
            "Lead Source",
            "Qualification",
            "status",
            "",
            "Initial Employee Remark",
            "Initial Called Date",
            "Source",
            "Initial Call Made",
        ]
        yellow_fill = PatternFill(start_color='fef2cb', end_color='fef2cb', fill_type='solid')  # Red fill

        # Define white fill pattern
        green_fill = PatternFill(start_color='c5e0b3', end_color='c5e0b3', fill_type='solid')  # green fill

        # Combine base headers and dynamically generated follow-up headers
        # headers = base_headers + sum([get_folloup_headers(calldetail) for calldetail in calldetails_data], [])
        headers = base_headers + sum([get_folloup_headers()], [])

        # Define font style for bold
        bold_font = Font(bold=True)
        

        # Write headers to the first row with appropriate fill applied
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header

            if cell.value:  # Check if cell value is not empty
                if col_idx <= len(base_headers):  # Apply red fill to base headers
                    cell.fill = yellow_fill
                else:  # Apply white fill to follow-up headers
                    cell.fill = green_fill

            # Apply bold font to each header cell
            cell.font = bold_font

        # Write headers to the first row only
        # Increase the height of the first row
        ws.row_dimensions[1].height = 30  # Adjust height as needed

        # # Remove color from empty cells
        # for row in ws.iter_rows():
        #     for cell in row:
        #         if cell.value is None:
        #             cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')  # White fill

        # ws.append(headers)
        # Set column width for each column


        for calldetail in calldetails_data:
            # Initialize empty lists to store folloup details
            folloup_remarks = []
            folloup_updated = []
            call_made_by = []

            # Retrieve related Folloup data
            folloup_data = Folloup.objects.filter(calldetails=calldetail)

            # Extract folloup details into separate lists
            for folloup in folloup_data:
                folloup_remarks.append(folloup.remark)
                folloup_updated.append(folloup.called_datetime.strftime("%d-%m-%Y"))  # Format date/time
                call_made_by.append(folloup.calls_made.name)  # Assuming 'name' is the relevant field

            # Extract Calldetails data
            row = []  # Create an empty row

            # Add lead details
            k = k + 1
            row.extend([
                k,
                calldetail.lead.control_no,
                calldetail.lead.date_time_added.strftime("%d-%m-%Y"),
                calldetail.lead.lead_given_date.strftime("%d-%m-%Y"),
                calldetail.lead.lead_no,
                calldetail.lead.name,
                calldetail.lead.course_type,
                calldetail.lead.course,
                calldetail.lead.phone_no,
                calldetail.lead.email,
                calldetail.lead.place,
                calldetail.lead.remark,
                calldetail.lead.source,
                calldetail.lead.degree,
                "denied",

                "",
                calldetail.emp_remark,
                calldetail.called_datetime.strftime("%d-%m-%Y"),
                calldetail.called_meadium,
                calldetail.calls_made.user_name,
            ])

            # Add folloup details with corresponding headings and empty columns
            for remark, updated, made_by in zip(folloup_remarks, folloup_updated, call_made_by):
                row.append("") # Add an empty column between each set of follow-up details
                row.append(made_by)
                row.append(updated)
                row.append(remark)


            # Write the combined row to the worksheet
            ws.append(row)


        ##### increase cell width ####
        for col in ws.columns:
            max_length = 0
            column = col[0].column  # Get the column index
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2  # Adjust the multiplication factor as needed
            ws.column_dimensions[get_column_letter(column)].width = adjusted_width




        # Define red fill pattern
        red_fill = PatternFill(start_color='ffcccc', end_color='ffcccc', fill_type='solid')  # Red fill

        # Get the index of the "Phone No" column
        phone_no_column_index = headers.index("Phone No") + 1  # Adding 1 because index starts from 1 in openpyxl

        # Iterate through each cell in the "Phone No" column
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=phone_no_column_index, max_col=phone_no_column_index):
            for cell in row:
                cell.fill = red_fill  # Apply red fill color to each cell


        # Define grape colour fill pattern
        grape_fill = PatternFill(start_color='ffccff', end_color='ffccff', fill_type='solid')  # grape fill

        # Get the index of the "Phone No" column
        phone_no_column_index = headers.index("Initial Call Made") + 1  # Adding 1 because index starts from 1 in openpyxl

        # Iterate through each cell in the "Phone No" column
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=phone_no_column_index, max_col=phone_no_column_index):
            for cell in row:
                cell.fill = grape_fill  # Apply red fill color to each cell

        
        # Iterate through each cell in the worksheet and set text alignment to center
        for row in ws.iter_rows():
            for cell in row:
                alignment = Alignment(horizontal='center', vertical='center')  # Center alignment
                cell.alignment = alignment

        # Define font style for Calibri
        calibri_font = Font(name='Calibri')
        # Iterate through each cell in the worksheet and set text to calibri_font
        for row in ws.iter_rows():
            for cell in row:
                # Apply Calibri font to each header cell
                cell.font = calibri_font


        # Create a response object
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=denied.xlsx'

        # Save the workbook to the response
        wb.save(response)
        return response
    else:
        return redirect('/')


def single_person_export_to_excel(request, id):
    if 'username' in request.session:
        k = 0

        calldetails_data = Calldetails.objects.filter(id=id)

        # Create a new workbook
        wb = Workbook()

        # Activate the first sheet
        ws = wb.active
        ws.title = "Need following Data"

        # def get_folloup_headers(calldetail):
        #     """
        #     Generates follow-up headers based on the number of follow-up entries for a calldetail.
        #     """
        #     folloup_count = calldetail.folloup_set.all().count()
        #     headers = []
        #     for _ in range(folloup_count):
        #         headers.extend(["", "Folloup Remark", "Folloup Updated", "Made By"])
        #     return headers


        def get_folloup_headers():
            ## print no of followups tables based on no of its won followup no
            calldetails_data1 = Calldetails.objects.get(id=id)
            highest_followups = calldetails_data1.no_of_followups
            highest_followups -= 1
            # folloup_count = calldetail.folloup_set.all().count()
            headers = []
            if highest_followups > 0:  # Add check to avoid empty headers if no follow-ups
                for i in range(highest_followups):
                    headers.extend(["", f"Follow-Up-{i+1}", "Date", "Remark"])
            return headers

        # Define base headers
        base_headers = [
            "SL.NO",
            "Control No",
            "Date Added",
            "Lead Given Date",
            "Lead No",
            "Name",
            "course type",
            "Course",
            "Phone No",
            "Email",
            "Place",
            "Lead Remark",
            "Lead Source",
            "Qualification",
            "status",
            "",
            "Initial Employee Remark",
            "Initial Called Date",
            "Source",
            "Initial Call Made",
        ]
        yellow_fill = PatternFill(start_color='fef2cb', end_color='fef2cb', fill_type='solid')  # Red fill

        # Define white fill pattern
        green_fill = PatternFill(start_color='c5e0b3', end_color='c5e0b3', fill_type='solid')  # green fill

        # Combine base headers and dynamically generated follow-up headers
        # headers = base_headers + sum([get_folloup_headers(calldetail) for calldetail in calldetails_data], [])
        headers = base_headers + sum([get_folloup_headers()], [])

        # Define font style for bold
        bold_font = Font(bold=True)
        

        # Write headers to the first row with appropriate fill applied
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header

            if cell.value:  # Check if cell value is not empty
                if col_idx <= len(base_headers):  # Apply red fill to base headers
                    cell.fill = yellow_fill
                else:  # Apply white fill to follow-up headers
                    cell.fill = green_fill

            # Apply bold font to each header cell
            cell.font = bold_font

        # Write headers to the first row only
        # Increase the height of the first row
        ws.row_dimensions[1].height = 30  # Adjust height as needed

        # # Remove color from empty cells
        # for row in ws.iter_rows():
        #     for cell in row:
        #         if cell.value is None:
        #             cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')  # White fill

        # ws.append(headers)
        # Set column width for each column


        for calldetail in calldetails_data:
            # Initialize empty lists to store folloup details
            folloup_remarks = []
            folloup_updated = []
            call_made_by = []

            # Retrieve related Folloup data
            folloup_data = Folloup.objects.filter(calldetails=calldetail)

            # Extract folloup details into separate lists
            for folloup in folloup_data:
                folloup_remarks.append(folloup.remark)
                folloup_updated.append(folloup.called_datetime.strftime("%d-%m-%Y"))  # Format date/time
                call_made_by.append(folloup.calls_made.name)  # Assuming 'name' is the relevant field

            # Extract Calldetails data
            row = []  # Create an empty row

            # Add lead details
            k = k + 1
            row.extend([
                k,
                calldetail.lead.control_no,
                calldetail.lead.date_time_added.strftime("%d-%m-%Y"),
                calldetail.lead.lead_given_date.strftime("%d-%m-%Y"),
                calldetail.lead.lead_no,
                calldetail.lead.name,
                calldetail.lead.course_type,
                calldetail.lead.course,
                calldetail.lead.phone_no,
                calldetail.lead.email,
                calldetail.lead.place,
                calldetail.lead.remark,
                calldetail.lead.source,
                calldetail.lead.degree,
                "need following",

                "",
                calldetail.emp_remark,
                calldetail.called_datetime.strftime("%d-%m-%Y"),
                calldetail.called_meadium,
                calldetail.calls_made.user_name,
            ])

            # Add folloup details with corresponding headings and empty columns
            for remark, updated, made_by in zip(folloup_remarks, folloup_updated, call_made_by):
                row.append("") # Add an empty column between each set of follow-up details
                row.append(made_by)
                row.append(updated)
                row.append(remark)


            # Write the combined row to the worksheet
            ws.append(row)


        ##### increase cell width ####
        for col in ws.columns:
            max_length = 0
            column = col[0].column  # Get the column index
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2  # Adjust the multiplication factor as needed
            ws.column_dimensions[get_column_letter(column)].width = adjusted_width




        # Define red fill pattern
        red_fill = PatternFill(start_color='ffcccc', end_color='ffcccc', fill_type='solid')  # Red fill

        # Get the index of the "Phone No" column
        phone_no_column_index = headers.index("Phone No") + 1  # Adding 1 because index starts from 1 in openpyxl

        # Iterate through each cell in the "Phone No" column
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=phone_no_column_index, max_col=phone_no_column_index):
            for cell in row:
                cell.fill = red_fill  # Apply red fill color to each cell


        # Define grape colour fill pattern
        grape_fill = PatternFill(start_color='ffccff', end_color='ffccff', fill_type='solid')  # grape fill

        # Get the index of the "Phone No" column
        phone_no_column_index = headers.index("Initial Call Made") + 1  # Adding 1 because index starts from 1 in openpyxl

        # Iterate through each cell in the "Phone No" column
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=phone_no_column_index, max_col=phone_no_column_index):
            for cell in row:
                cell.fill = grape_fill  # Apply red fill color to each cell

        
        # Iterate through each cell in the worksheet and set text alignment to center
        for row in ws.iter_rows():
            for cell in row:
                alignment = Alignment(horizontal='center', vertical='center')  # Center alignment
                cell.alignment = alignment

        # Define font style for Calibri
        calibri_font = Font(name='Calibri')
        # Iterate through each cell in the worksheet and set text to calibri_font
        for row in ws.iter_rows():
            for cell in row:
                # Apply Calibri font to each header cell
                cell.font = calibri_font


        # Create a response object
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=needfollowing.xlsx'

        

        print(calldetails_data)
        for i in calldetails_data:
            filename = f"{i.lead.control_no}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'


        # Save the workbook to the response
        wb.save(response)
        return response
    else:
        return redirect('/')

def edit(request, id):
    if 'username' in request.session:
        if request.method == 'POST':
            print("post part")
            name  = request.POST['name1']
            course = request.POST['course']
            phone_no = request.POST['phone_no']
            email = request.POST['email']
            place = request.POST['place']
            source = request.POST['source']
            degree = request.POST['degree']
            remark = request.POST['remark']
            try:
                status = request.POST['status']
            except:
                status = 0


            course_types = request.POST.get('coursemode')
            if course_types == "Not mentioned":
                course_type = ""
            else:
                course_type = course_types

            Lead.objects.filter(id=id).update(name=name, course=course, phone_no=phone_no, email=email, place=place, source=source, degree=degree, course_type=course_type, status=status, remark=remark)
            print("post part")
            # Redirect to the home page with refresh parameter
            return redirect('/')
        data = Lead.objects.filter(id=id)
        coursedata = Courses.objects.all()
        return render(request,'edit.html',{'data':data,'coursedata':coursedata})
    else:
        return redirect('/')
    
def update_priority(request):
    print("%%%^^^&&&&&&&&&&")
    # if request.method == 'POST' and request.is_ajax():
    if request.method == 'POST':
        priority = request.POST.get('priority')
        person_id = request.POST.get('person_id')

        print(f"person_id- {person_id}")
        print(f"priority- {priority}")
        
        # Update the priority in your model
        try:
            person = Lead.objects.get(id=person_id)
            print(person.control_no)
            person.priority = priority
            person.save()
            return JsonResponse({'message': 'Priority updated successfully'}, status=200)
        except Lead.DoesNotExist:
            return JsonResponse({'message': 'Person not found'}, status=404)

    return JsonResponse({'error': 'Invalid request'}, status=400)